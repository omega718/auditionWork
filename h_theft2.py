import pandas as pd,os
from openpyxl import load_workbook as lwb
from openpyxl.chart import BarChart, Reference
from collections import Counter

fileName="chart_theft.xlsx"  
# #從政府資料開放平台抓取台北市住宅竊盜點位資訊之csv檔案
url="https://data.taipei/api/getDatasetInfo/downloadResource?id=68785231-d6c5-47a1-b001-77eec70bec02&rid=34a4a431-f04d-474a-8e72-8d3f586db3df"
pRead = pd.read_csv(url,encoding="cp950",index_col=0)
#pRead = pd.read_csv("thief.csv",encoding="cp950",index_col=0)

#從地址欄取出區資料
areaList,roadList,hoursList=[],[],[]
address = pRead.iloc[:,3] #地址
hours = pRead.iloc[:,2] #時間
for line in address:
    #台北市以外的區不處理
    if line[0:3]=="台北市":
        #第一個字從0開始
        p=line.find("區")
        line1 = line[3:p+1]
    else:
        line1=""
    areaList.append(line1)
    if line.find("區")!=-1:
        p1=line.find("區")
        p2=line.find("路")
        p3=line.find("道")
        p4=line.find("街")
        x=p2 if p2!=-1 else (p3 if p3!=-1 else (p4 if p4!=-1 else -1))
        line1=line[p1+1:x+1]
    else:
        line1=""
    roadList.append(line1)
for line in hours:
    hoursList.append(line)
#Counter以區進行計數統計
#most_common函數返回Counter中次數最多的N個元素，如果N沒有提供或者是None，
#那麼就會返回所有元素。
datas1,idx1=[],[]
areaCounter=Counter(areaList).most_common()
for (k,l) in areaCounter:
    idx1.append(k)
    datas1.append(l)
# indexs.append("總計:")
# datas.append(len(areaList))
a=list(pRead.iloc[:,1])
b=list(pRead.iloc[:,2])
c=list(pRead.iloc[:,3])
d=areaList
e=roadList
df=pd.DataFrame({"發生(現)日期":a,"發生時段":b,"發生(現)地點":c,"區名":d,"路名":e})
columns=["計數-區"]
df1 = pd.DataFrame(datas1, columns=columns,  index=idx1)
datas2,idx2=[],[]
roadCounter=Counter(roadList).most_common(10)
for (k,l) in roadCounter:
    idx2.append(k)
    datas2.append(l)
columns=["計數-路"]
df2 = pd.DataFrame(datas2, columns=columns,  index=idx2)
datas3,idx3=[],[]
hourCounter=Counter(hoursList).most_common()
for (k,l) in hourCounter:
    idx3.append(k)
    datas3.append(l)
columns=["計數-發生時段"]
df3 = pd.DataFrame(datas3, columns=columns,  index=idx3)

writer = pd.ExcelWriter(fileName) #writer寫到目的xlsx檔
df.to_excel(writer, "原始資料",startrow=0, startcol=0,index=False)
df1.to_excel(writer, "DataBy區",startrow=0, startcol=0)
df2.to_excel(writer, "DataBy路",startrow=0, startcol=0)
df3.to_excel(writer, "DataBy時段",startrow=0, startcol=0)
writer.save()
pRead = pd.read_excel(fileName,"原始資料")
idx3.sort() #將時段排序 
idx1.sort() #將區名排序
idx1.remove("")#除去區名空白

datas4=[]
for hour in idx3:
    tempHour=[]
    df_hour = pRead[pRead.發生時段 == hour]
    #print(df_hour)
    tempArea=list(df_hour.iloc[:,3])#groupby時段之下的區名
    for area in idx1:
        tempHour.append(tempArea.count(area))
    datas4.append(tempHour)

columns=idx3
fmtdict={}

for i in range(len(idx3)):
    fmtdict[idx3[i]]=datas4[i]
print(fmtdict)     
df4 = pd.DataFrame(fmtdict, columns=idx3,  index=idx1)
df4.to_excel(writer, "By區和時段",startrow=0, startcol=0)
writer.save()

#畫圖
wbook = lwb(fileName) #openpyxl.workload不接受.xls
chart1 = BarChart()
chart1.type = "col"
chart1.style = 5
chart1.title = "合計by區"
chart1.height=10
chart1.width=20
wsheet=wbook["DataBy區"]
data = Reference(wsheet, min_col=wsheet.min_column+1, min_row=wsheet.min_row, max_row=wsheet.max_row-1)
cats = Reference(wsheet, min_col=wsheet.min_column, min_row=wsheet.min_row+1, max_row=wsheet.max_row-1)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)

wsheet.add_chart(chart1, "C2")
wsheet=wbook["DataBy路"]
chart2=BarChart()
chart2.style = 6
chart2.title = "合計by路"
data = Reference(wsheet, min_col=wsheet.min_column+1, min_row=wsheet.min_row, max_row=wsheet.max_row)
cats = Reference(wsheet, min_col=wsheet.min_column, min_row=wsheet.min_row+1, max_row=wsheet.max_row)
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(cats)
chart2.height=10
chart2.width=20
wsheet.add_chart(chart2, "C2")

wsheet=wbook["DataBy時段"]
chart3=BarChart()
chart3.style = 7
chart3.height=10
chart3.width=20
chart3.title = "合計by時段"
data = Reference(wsheet, min_col=wsheet.min_column+1, min_row=wsheet.min_row, max_row=wsheet.max_row)
cats = Reference(wsheet, min_col=wsheet.min_column, min_row=wsheet.min_row+1, max_row=wsheet.max_row)
chart3.add_data(data, titles_from_data=True)
chart3.set_categories(cats)
wsheet.add_chart(chart3, "C2")

wsheet=wbook["By區和時段"]
chart4=BarChart()
chart4.styles=range(0,len(idx3)+1)
chart4.height=14
chart4.width=22
chart4.title="計數-區和時段"
chart4.type = "col"
chart4.grouping = "percentStacked" #百分比堆疊圖
chart4.overlap = 100
data = Reference(wsheet, min_col=wsheet.min_column+1,max_col=wsheet.max_column,min_row=wsheet.min_row, max_row=wsheet.max_row)
cats = Reference(wsheet, min_col=wsheet.min_column, min_row=wsheet.min_row+1, max_row=wsheet.max_row)
chart4.add_data(data, titles_from_data=True)
chart4.set_categories(cats)
wsheet.add_chart(chart4, "B14")
wbook.save(fileName)
path=os.getcwd() + "\\" +fileName
print("已產出檔案:",path,",結束")